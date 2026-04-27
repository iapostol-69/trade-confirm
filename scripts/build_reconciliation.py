"""
build_reconciliation.py
-----------------------
Helper script for the trade-confirm skill.
Builds a formatted Excel reconciliation file from matched PO vs OC data.

Usage:
    python build_reconciliation.py <data.json> <output.xlsx>

Input JSON format:
{
  "order_ref": "2025-1933",
  "mismatches": [
    {
      "order_our_code":      "071203",   // our internal product code from the order file; null if missing
      "order_supplier_code": "CUST-001", // supplier product code from the order file; null if missing
      "order_desc":          "Stainless pipe 50mm",
      "order_qty":           10,
      "order_price":         25.50,      // unit price; divide line total by qty if needed
      "conf_our_code":       "071203",   // our internal product code from the confirmation; null if missing
      "conf_supplier_code":  "CUST-001", // supplier product code from the confirmation; null if missing
      "conf_desc":           "SS Pipe 50mm",
      "conf_qty":            8,
      "conf_price":          27.00,
      "match_type":          "Our Code", // "Our Code", "Supplier Code", "Description", or "Unmatched"
      "desc_mismatch":       false       // true when matched by code but descriptions suggest different product
    }
  ],
  "matches": [
    { "...same fields..." }
  ],
  "extra_in_confirmation": [
    {
      "conf_our_code":      "123544",
      "conf_supplier_code": "ART.999",
      "conf_desc":          "Extra item not in order",
      "conf_qty":           5,
      "conf_price":         12.00
    }
  ]
}
"""

import sys
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- Colors ----------
RED_HEADER_FILL   = PatternFill("solid", fgColor="FFCC3333")
GREEN_HEADER_FILL = PatternFill("solid", fgColor="FF2E7D32")
ROW_MISMATCH_FILL = PatternFill("solid", fgColor="FFFFCCCC")
ROW_MATCH_FILL    = PatternFill("solid", fgColor="FFCCFFCC")
CELL_DIFF_FILL    = PatternFill("solid", fgColor="FFFF6666")
UNMATCHED_FILL    = PatternFill("solid", fgColor="FFFFF0CC")
SUMMARY_FILL      = PatternFill("solid", fgColor="FFEEF2FF")
EXTRA_FILL        = PatternFill("solid", fgColor="FFFFE4B5")

WHITE_BOLD   = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
SUMMARY_FONT = Font(name="Arial", size=10, bold=True, color="FF1A237E")
DIFF_FONT    = Font(name="Arial", size=10, bold=True, color="FF880000")

# 11 columns: primary code, secondary code, description, qty, price — for each side, plus match type
HEADERS = [
    "Our Code (Order)", "Supplier Code (Order)", "Description (Order)", "Qty (Order)", "Price (Order)",
    "Our Code (Confirm.)", "Supplier Code (Confirm.)", "Description (Confirm.)", "Qty (Confirm.)", "Price (Confirm.)",
    "Match Type"
]

# Column positions (1-based) for field types — update here if HEADERS ever changes
COL_DESC_ORDER   = 3
COL_QTY_ORDER    = 4
COL_PRICE_ORDER  = 5
COL_DESC_CONF    = 8
COL_QTY_CONF     = 9
COL_PRICE_CONF   = 10

THIN = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _write_header(ws, row, fill):
    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = WHITE_BOLD
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _write_data_row(ws, row, item, row_fill,
                    highlight_qty=False, highlight_price=False,
                    highlight_desc=False, extra=False):
    values = [
        item.get("order_our_code",      "") or "",
        item.get("order_supplier_code", "") or "",
        item.get("order_desc",          ""),
        item.get("order_qty",           ""),
        item.get("order_price",         ""),
        item.get("conf_our_code",       "") or "",
        item.get("conf_supplier_code",  "") or "",
        item.get("conf_desc",           ""),
        item.get("conf_qty",            ""),
        item.get("conf_price",          ""),
        item.get("match_type",          ""),
    ]
    diff_cols = set()
    if highlight_desc:
        diff_cols |= {COL_DESC_ORDER, COL_DESC_CONF}
    if highlight_qty:
        diff_cols |= {COL_QTY_ORDER, COL_QTY_CONF}
    if highlight_price:
        diff_cols |= {COL_PRICE_ORDER, COL_PRICE_CONF}

    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")
        if col in diff_cols:
            cell.fill = CELL_DIFF_FILL
            cell.font = DIFF_FONT
        else:
            cell.fill = row_fill
            cell.font = NORMAL

    # Number formatting
    for col in (COL_QTY_ORDER, COL_QTY_CONF):
        ws.cell(row=row, column=col).number_format = "#,##0.##"
    for col in (COL_PRICE_ORDER, COL_PRICE_CONF):
        ws.cell(row=row, column=col).number_format = "#,##0.00"


def _auto_fit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        adjusted = min(max(max_len + 2, 12), 50)
        ws.column_dimensions[col_letter].width = adjusted


def build(data: dict, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reconciliation"
    ws.sheet_view.showGridLines = True

    mismatches = data.get("mismatches", [])
    matches    = data.get("matches", [])
    extras     = data.get("extra_in_confirmation", [])
    order_ref  = data.get("order_ref", "")

    total_order_items = len(mismatches) + len(matches)
    n_mismatches  = len(mismatches)
    n_matches     = len(matches)
    n_unmatched   = sum(1 for m in mismatches if m.get("match_type") == "Unmatched")
    n_extra       = len(extras)

    # ---- Row 1: Summary ----
    summary_text = (
        f"Order ref: {order_ref}   |   "
        f"Order items: {total_order_items}   |   "
        f"Mismatches: {n_mismatches - n_unmatched}   |   "
        f"Unmatched: {n_unmatched}   |   "
        f"Matches: {n_matches}"
        + (f"   |   Extra in confirmation: {n_extra}" if n_extra else "")
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    summary_cell = ws.cell(row=1, column=1, value=summary_text)
    summary_cell.font = SUMMARY_FONT
    summary_cell.fill = SUMMARY_FILL
    summary_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    current_row = 2

    # ---- Section 1: Mismatches ----
    if mismatches:
        _write_header(ws, current_row, RED_HEADER_FILL)
        ws.row_dimensions[current_row].height = 28
        current_row += 1

        for item in mismatches:
            is_unmatched = item.get("match_type") == "Unmatched"
            row_fill = UNMATCHED_FILL if is_unmatched else ROW_MISMATCH_FILL

            try:
                qty_diff = float(item.get("order_qty") or 0) != float(item.get("conf_qty") or 0)
            except (TypeError, ValueError):
                qty_diff = item.get("order_qty") != item.get("conf_qty")
            try:
                price_diff = abs(float(item.get("order_price") or 0) - float(item.get("conf_price") or 0)) >= 0.01
            except (TypeError, ValueError):
                price_diff = item.get("order_price") != item.get("conf_price")
            desc_diff = bool(item.get("desc_mismatch")) and not is_unmatched

            _write_data_row(ws, current_row, item, row_fill,
                            highlight_qty=qty_diff and not is_unmatched,
                            highlight_price=price_diff and not is_unmatched,
                            highlight_desc=desc_diff)
            ws.row_dimensions[current_row].height = 18
            current_row += 1

    # ---- Blank separator ----
    current_row += 1

    # ---- Section 2: Matches ----
    if matches:
        _write_header(ws, current_row, GREEN_HEADER_FILL)
        ws.row_dimensions[current_row].height = 28
        current_row += 1

        for item in matches:
            _write_data_row(ws, current_row, item, ROW_MATCH_FILL)
            ws.row_dimensions[current_row].height = 18
            current_row += 1

    # ---- Section 3: Extra items in confirmation (not in order) ----
    if extras:
        current_row += 1
        extra_header_fill = PatternFill("solid", fgColor="E65100")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(HEADERS))
        label = ws.cell(row=current_row, column=1,
                        value="Items in Confirmation NOT found in the Order")
        label.font = WHITE_BOLD
        label.fill = extra_header_fill
        label.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1

        _write_header(ws, current_row, extra_header_fill)
        current_row += 1
        for item in extras:
            extra_item = {
                "order_our_code":      "",
                "order_supplier_code": "",
                "order_desc":          "",
                "order_qty":           "",
                "order_price":         "",
                "conf_our_code":       item.get("conf_our_code",      "") or "",
                "conf_supplier_code":  item.get("conf_supplier_code", "") or "",
                "conf_desc":           item.get("conf_desc",          ""),
                "conf_qty":            item.get("conf_qty",           ""),
                "conf_price":          item.get("conf_price",         ""),
                "match_type":          "Extra",
            }
            _write_data_row(ws, current_row, extra_item, EXTRA_FILL)
            current_row += 1

    # ---- Auto-fit and freeze ----
    _auto_fit_columns(ws)
    ws.freeze_panes = "A3"  # Keep summary + first header visible

    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python build_reconciliation.py <data.json> <output.xlsx>")
        sys.exit(1)
    with open(sys.argv[1]) as f:
        data = json.load(f)
    build(data, sys.argv[2])
